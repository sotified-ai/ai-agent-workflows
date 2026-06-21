import os
import pandas as pd
from typing import Dict, Any, List

class SpreadsheetExporter:
    """
    Exports the generated test engineering artifacts (Requirements, Scenarios, 
    Test Cases, and Traceability) to a production-grade multi-sheet Excel file.
    Follows Gumloop spreadsheet-output rules strictly:
    - Unique column headers in row 1, no titles/metadata, no blank rows.
    """
    @staticmethod
    def export_to_excel(pipeline_output: Dict[str, Any], filepath: str) -> str:
        # Create separate DataFrames for each Sheet

        # 1. Requirements Sheet
        req_rows = []
        for r in pipeline_output["requirements"]:
            req_rows.append({
                "Requirement ID": r["id"],
                "Title": r["title"],
                "Description": r["description"],
                "Criticality": r["criticality"],
                "Compliance Standards": ", ".join(r.get("compliance_tags", []))
            })
        df_req = pd.DataFrame(req_rows)

        # 2. Scenarios Sheet
        sc_rows = []
        for s in pipeline_output["scenarios"]:
            sc_rows.append({
                "Scenario ID": s["id"],
                "Requirement ID": s["requirement_id"],
                "Title": s["title"],
                "Description": s["description"],
                "Type": s["type"]
            })
        df_sc = pd.DataFrame(sc_rows)

        # 3. Test Cases Sheet
        tc_rows = []
        for t in pipeline_output["test_cases"]:
            tc_rows.append({
                "Test Case ID": t["id"],
                "Scenario ID": t["scenario_id"],
                "Requirement ID": t["requirement_id"],
                "Test Case Name": t["name"],
                "Preconditions": t["preconditions"],
                "Execution Steps": t["steps"],
                "Expected Result": t["expected_result"],
                "Type": t["type"],
                "Automation Candidate": t["automation_candidate"]
            })
        df_tc = pd.DataFrame(tc_rows)

        # 4. Traceability Summary Sheet
        trace_data = pipeline_output["traceability"]
        trace_rows = []
        for req_id, sc_ids in trace_data.get("req_to_scenarios_map", {}).items():
            # Find related test cases
            tc_ids = []
            for tc in pipeline_output["test_cases"]:
                if tc["requirement_id"] == req_id:
                    tc_ids.append(tc["id"])

            trace_rows.append({
                "Requirement ID": req_id,
                "Associated Scenarios": ", ".join(sc_ids),
                "Associated Test Cases": ", ".join(tc_ids),
                "Coverage Status": "Fully Covered" if tc_ids else "Gap Identified"
            })
        df_trace = pd.DataFrame(trace_rows)

        from pathlib import Path
        filepath = Path(filepath)
        filepath.parent.mkdir(parents=True, exist_ok=True)

        # Write to Excel with multiple sheets using ExcelWriter
        with pd.ExcelWriter(str(filepath), engine='openpyxl') as writer:
            df_req.to_excel(writer, sheet_name="Requirements", index=False)
            df_sc.to_excel(writer, sheet_name="Test Scenarios", index=False)
            df_tc.to_excel(writer, sheet_name="Test Cases", index=False)
            df_trace.to_excel(writer, sheet_name="Traceability Matrix", index=False)

            # Optional: Format spreadsheet using openpyxl for clean styling
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            workbook = writer.book
            header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid") # Charcoal Theme
            cell_font = Font(name="Segoe UI", size=10)
            thin_border = Border(
                left=Side(style='thin', color='E5E7EB'),
                right=Side(style='thin', color='E5E7EB'),
                top=Side(style='thin', color='E5E7EB'),
                bottom=Side(style='thin', color='E5E7EB')
            )

            for sheet_name in workbook.sheetnames:
                ws = workbook[sheet_name]
                ws.auto_filter.ref = ws.dimensions
                ws.row_dimensions[1].height = 26

                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

                for row in range(2, ws.max_row + 1):
                    ws.row_dimensions[row].height = 20
                    for col in range(1, ws.max_column + 1):
                        cell = ws.cell(row=row, column=col)
                        cell.font = cell_font
                        cell.border = thin_border
                        cell.alignment = Alignment(vertical="center", wrap_text=True)

                for col in ws.columns:
                    max_len = 0
                    col_letter = col[0].column_letter
                    for cell in col:
                        val = str(cell.value or '')
                        if '\n' in val:
                            lines = val.split('\n')
                            max_len = max(max_len, max(len(l) for l in lines))
                        else:
                            max_len = max(max_len, len(val))
                    ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 40)

            workbook.save(str(filepath))
        return str(filepath)
